from openpyxl import load_workbook
from datetime import date
import os
BASE_PATH=os.path.dirname(os.path.abspath(__file__))
file_path=os.path.join(BASE_PATH,'data','invoice_data.xlsx')
def get_next_invoice_number():
    wb=load_workbook(file_path,keep_vba=True)
    current_page=wb['Invoice data']
    if current_page.max_row ==1:
        print("Data doesn't exist in this file !")
        return 123456
    else:
        last_row_invoice=f"A{current_page.max_row}"
        last_row_invoice_number=current_page[last_row_invoice].value
        new_invoice=last_row_invoice_number+1
        return new_invoice


def save_invoice(build_row):
    wb=load_workbook(file_path)
    current_page=wb['Invoice data']
    for item in build_row['items']:
        building_row = [
                build_row['invoice_number'],build_row['date'],build_row['company_name'],
                item['item_code'],item['description'],item['item_price'],item['amount'],
                build_row['salesman'],build_row['po_no'],"Sales/Service Invoice",
                build_row["shipping"] ]
        current_page.append(building_row)
    wb.save(file_path)


#OPTIONAL(For CL-Interface)
# def main():
#     while True:
#         invoice_number=input("Enter Invoice Number: ")
#         if invoice_number=='':
#             invoice_number=get_next_invoice_number()
#         else:
#             invoice_number=int(invoice_number)
#         date_today=input('Date: ')
#         if date_today=="":
#             date_today=date.today()
#         company_name=input("Company Name: ")
#         salesman=input("Salesman: ")
#         po_no=input("Party GSTIN: ").upper()
#         shipping=input("From which country is the order Shipped From : ")
#         if shipping=="":
#             shipping='INDIA'
#         number_of_items=int(input('Number of Items on this Invoice: '))
#         if number_of_items=='':
#             number_of_items=int(1)
#         elif number_of_items<0:
#             print('Wrong input!')
#         else:
#             number_of_items=int(number_of_items)
#             new_items=[]
#             for number in range(1,number_of_items+1):
#                 print(f"------For Item {number}--------")
#                 item_code=int(input("Item Code : "))
#                 description=input("Description : ")
#                 item_price=int(input("Item Price : "))
#                 amount=int(input("Enter Amount : "))
#                 new_item={"item_code": item_code, "description": description, "item_price": item_price, "amount": amount}
#                 new_items.append(new_item)
#             user_data = {
#             "invoice_number": invoice_number,
#             "date": date_today,
#             "company_name": company_name,
#             "salesman": salesman,
#             "po_no": po_no,
#             "shipping": shipping,
#             "items": new_items
#         }
#             print('I got the Full User data-> ',user_data)
#             save_invoice(user_data)
#             terminate=input("Terminate Process? y/n : ").lower()
#             if terminate=='y':
#                 break

# main()

# def user_invoice_input():
#     try:
#         invoice=input("Enter invoice Number: ")
#         if invoice=='':
#             print(get_next_invoice_number())
#         else:
#             invoice=int(invoice)
#             print(invoice) #works fine
#             # print(type(invoice)) #returns int class
#     except ValueError as E:
#             print('The error: ',E)
# user_invoice_input()
