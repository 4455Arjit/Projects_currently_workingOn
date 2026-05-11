import os
from openpyxl import Workbook
from openpyxl import load_workbook
def filee_path():
    utils_path=os.path.dirname(os.path.abspath(__file__))
    BASE_PATH=os.path.dirname(utils_path)
    file_path=os.path.join(BASE_PATH,'data','invoice_data.xlsx')
    return file_path
file_path=filee_path()
def working():
    try:
        if os.path.exists(file_path):
            load_wb=load_workbook(file_path)
            sheet_active=load_wb['Invoice data']
            # headers=['Invoice No.','Item code/Qty','Description','Date','Item Value','AMOUNT','Sale/Service-Cat','Party GSTIN','Sold To:(Address 4 lines)']
            headers=['Invoice','Date','Company Name','Item Code','Description','Item price','Amount','salesmen','PO no','Notfy','Shipping']
            if sheet_active['A1'].value=='Invoice':
                print("File exists and contains all necessary values! Don't do anything now . ")
            else:
                for col,header in enumerate(headers, start=1):
                    sheet_active.cell(row=1,column=col,value=header)
                load_wb.save(file_path)
        else:
            wb = Workbook()
            page_1 = wb.create_sheet('Invoice data',0)
            wb.remove(wb['Sheet'])
            wb.save(file_path)
    except Exception as e:
        print("We ran into an error: ",e)
working()